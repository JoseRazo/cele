import xlwt
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.views.generic.list import ListView
from datetime import date as fecha_actual
from django.db.models import Count, Q, Value as V, Avg, Sum
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.utils import timezone
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from utils.meses import MONTHS_SPANISH

from gestion_escolar.models import (
    Alumno, 
    Curso as CursoCELE,
    CursoAlumno,
    Profesor,
    Periodo as PeriodoCELE
    )
from edcon.models import (
    Curso as CursoEDCON,
    Estudiante, 
    CursoEstudiante,
    Instructor,
    Periodo as PeriodoEDCON
    )
from .models import (
    Encuesta,
    Seccion,
    Pregunta,
    EncuestaAlumno, 
    EncuestaEstudiante, 
    RegistrosCELE, 
    RegistrosEDCON
    )

# Create your views here.
@login_required
def realizar_encuesta(request, curso_id, type):
    usuario = request.user
    grupos = request.user.groups.all()

    if type == 'AC':
        selcurso = CursoAlumno.objects.get(pk=curso_id)
    elif type == 'AE':
        selcurso = CursoEstudiante.objects.get(pk=curso_id)

    if usuario.is_superuser == 1:
            usuario_log = usuario
            status = 'admin'

    for grupo in grupos:
        if grupo.name == 'Alumnos CELE':
            usuario_log = Alumno.objects.get(username=usuario.username)
            selcurso = CursoAlumno.objects.get(pk=curso_id)
            log_alumno = str(selcurso.alumno)
            status = 'alumno'
        elif grupo.name == 'Estudiantes EDCON':
            usuario_log = Estudiante.objects.get(username=usuario.username)
            selcurso = CursoEstudiante.objects.get(pk=curso_id)
            log_alumno = str(selcurso.estudiante)
            status = 'estudiante'
        elif grupo.name == 'Administradores CELE' or grupo.name == 'Administradores EDCON':
            usuario_log = request.user
            status = 'admin'
        else:
            usuario_log = request.user
            return render(request, 'certificados/404.html')
    
    try:
        if status == 'alumno' or type == 'AC':
            encuesta_res = EncuestaAlumno.objects.filter(curso_alumno_id=curso_id)
            encuesta_static = RegistrosCELE.objects.filter(curso_alumno_id=curso_id)
        elif status == 'estudiante' or type == 'AE':
            encuesta_res = EncuestaEstudiante.objects.filter(curso_alumno_id=curso_id)
            encuesta_static = RegistrosEDCON.objects.filter(curso_alumno_id=curso_id)

    except ObjectDoesNotExist:
        pass

    encuesta = Encuesta.objects.get(nombre='REVIN025')
    secciones = Seccion.objects.filter(encuesta=encuesta)
    preguntas = encuesta.pregunta_set.filter(activo=True)
       
    usuario = str(request.user)

    if request.method == 'POST':
        if str(encuesta_res) == '<QuerySet []>':
            if status == 'alumno':
                registro = RegistrosCELE(
                    usuario=request.user,
                    curso_alumno=selcurso,
                    encuesta=encuesta
                )
                registro.save()
            elif status == 'estudiante':
                registro = RegistrosEDCON(
                    usuario=request.user,
                    curso_alumno=selcurso,
                    encuesta=encuesta
                )
                registro.save()

            for pregunta in preguntas:
                respuesta_texto = request.POST.get(f"pregunta_{pregunta.id}")

                if respuesta_texto:
                    if status == 'alumno':
                        solicitud_encuesta = EncuestaAlumno(
                            curso_alumno=selcurso,
                            registro=registro,
                            encuesta=encuesta,
                            pregunta=pregunta,
                            texto_respuesta=respuesta_texto
                        )
                        solicitud_encuesta.save()
                    elif status == 'estudiante':
                        solicitud_encuesta = EncuestaEstudiante(
                            curso_alumno=selcurso,
                            registro=registro,
                            encuesta=encuesta,
                            pregunta=pregunta,
                            texto_respuesta=respuesta_texto
                        )
                        solicitud_encuesta.save()

    context = {
        'selcurso': selcurso,
        'usuario_log': usuario_log, 
        'encuesta': encuesta, 
        'secciones': secciones, 
        'preguntas': preguntas, 
        'encuesta_res': encuesta_res, 
        'encuesta_static': encuesta_static, 
        'status': status
    }

    if type == 'AC' or type == 'AE':
        return render(request, 'encuestas/realizar_encuesta.html', context)
    else:
        if log_alumno == usuario:
            # El curso de alumno no pertenece al usuario logueado, mostrar un mensaje de error o redirigir a otra página
            return render(request, 'encuestas/realizar_encuesta.html', context)
        else:
            return render(request, 'certificados/curso_no_autorizado.html',
                        {'usuario_log': usuario_log})

# @login_required
# def reportgen(request, type):
#     try:
#         grupos = request.user.groups.all()
#         for grupo in grupos :
#             if grupo.name == 'Administradores CELE' or grupo.name == 'Administradores EDCON' or request.user.is_superuser == 1:
#                 status = 'admin'
#             else:
#                 return render(request, 'certificados/404.html')
        
#         filename = ''
#         wb = xlwt.Workbook(encoding='utf-8')
#         ws = wb.add_sheet('Reporte')

#         row_num = 0

#         font_style = xlwt.XFStyle()
#         font_style.font.bold = True

#         selcurso = request.GET.get('curso')
#         selprofe = request.GET.get('profe')
#         selperiodo = request.GET.get('periodo')

#         if not selcurso and not selprofe and not selperiodo:
#             messages.error(request, 'Selecciona al menos 1 filtro.')
#             if type == "AC":
#                 return redirect('/reportes-cele')
#             elif type == "AE":
#                 return redirect('/reportes-edcon')

#         filtros = {}

#         if selcurso:
#             filtros['curso_alumno__curso__nombre'] = selcurso
#             filename += '_' + str(selcurso)
#         if selprofe:
#             filtros['nombres'] = selprofe
#             filename += '_' + str(selprofe)
#         if selperiodo:
#             filtros['curso_alumno__periodo'] = selperiodo
#             filename += '_' + str(selperiodo)

#         if type == "AC":
#             selreporte = RegistrosCELE.objects.annotate(nombres=Concat('curso_alumno__profesor__nombre', V(' '),'curso_alumno__profesor__apellido_paterno', V(' '), 'curso_alumno__profesor__apellido_materno')).filter(**filtros)
#             reporte_type = "CELE"
#         elif type == "AE":
#             selreporte = RegistrosEDCON.objects.annotate(nombres=Concat('curso_alumno__instructor__nombre', V(' '),'curso_alumno__instructor__apellido_paterno', V(' '), 'curso_alumno__instructor__apellido_materno')).filter(**filtros)
#             reporte_type = "EDCON"

#         encuesta = selreporte[0].encuesta
#         get_preguntas = encuesta.pregunta_set.filter(activo=True)
#         preguntas = []
#         for pregunta in get_preguntas:
#             preguntas.append(pregunta.texto_pregunta)

#         datos = []

#         columns = ["Curso", "Alumno/Estudiante", "Profesor", "Periodo"]
#         columns.extend(preguntas)


#         for col_num, col in enumerate(columns):
#             ws.write(row_num, col_num, col, font_style)
#             cwidth = ws.col(col_num).width
#             if (len(col)*367) > cwidth:  
#                 ws.col(col_num).width = len(col)*300
        
        
#         font_style = xlwt.XFStyle()

#         for index, obj in enumerate(selreporte):
#             if type == "AC":
#                 encuestas = EncuestaAlumno.objects.annotate(nombres=Concat('curso_alumno__profesor__nombre', V(' '),'curso_alumno__profesor__apellido_paterno', V(' '), 'curso_alumno__profesor__apellido_materno')).filter(**filtros)
#             elif type == "AE":
#                 encuestas = EncuestaEstudiante.objects.annotate(nombres=Concat('curso_alumno__instructor__nombre', V(' '),'curso_alumno__instructor__apellido_paterno', V(' '), 'curso_alumno__instructor__apellido_materno')).filter(**filtros)

#             respuestas = []
#             for enc in encuestas:
#                 if type == "AC":
#                     if enc.curso_alumno.curso == selreporte[index].curso_alumno.curso and enc.curso_alumno.alumno == selreporte[index].curso_alumno.alumno:
#                         respuestas.append(enc.texto_respuesta)
#                 elif type == "AE":
#                     if enc.curso_alumno.curso == selreporte[index].curso_alumno.curso and enc.curso_alumno.estudiante == selreporte[index].curso_alumno.estudiante:
#                         respuestas.append(enc.texto_respuesta)

#             datos = [str(obj.curso_alumno)]
#             datos.append(str(obj.usuario))
#             if type == "AC":
#                 datos.append(str(obj.curso_alumno.profesor))
#             elif type == "AE":
#                 datos.append(str(obj.curso_alumno.instructor))
#             datos.append(str(obj.curso_alumno.periodo))
#             datos.extend(respuestas)

#             for col_num, dato in enumerate(datos):
#                 ws.write(index + 1, col_num, dato, font_style)


#         today = fecha_actual.today()
#         filename = 'attachment; filename="Reporte' + reporte_type + filename + '_' + str(today) + '.xls"'
#         response = HttpResponse(
#             content_type='application/ms-excel',
#             headers={"Content-Disposition": filename}
#             )

#         wb.save(response)

#         return response
#     except:
#         messages.error(request, 'No existen registros que cumplan con los filtros seleccionados.')
#         if type == "AC":
#             return redirect('/reportes-cele')
#         elif type == "AE":
#             return redirect('/reportes-edcon')
        
@login_required
def reportgenrevin026(request, type):
    if not request.GET.get('profe') or not request.GET.get('curso') or not request.GET.get('periodo'):
        messages.error(request, 'Por favor, selecciona todos los filtros.')
        if type == "AC":
            return redirect('/reportes-cele')
        elif type == "AE":
            return redirect('/reportes-edcon')
    
    if type == "AC":
        selcurso = CursoCELE.objects.get(pk=request.GET.get('curso'))
        selprofe = Profesor.objects.get(pk=request.GET.get('profe'))
        selperiodo = PeriodoCELE.objects.get(pk=request.GET.get('periodo'))
    elif type == "AE":
        selcurso = CursoEDCON.objects.get(pk=request.GET.get('curso'))
        selprofe = Instructor.objects.get(pk=request.GET.get('profe'))
        selperiodo = PeriodoEDCON.objects.get(pk=request.GET.get('periodo'))

    inicio_dia_periodo = selperiodo.fecha_inicio.strftime('%d')
    inicio_mes_periodo = selperiodo.fecha_inicio.strftime('%B')
    inicio_anio_periodo = selperiodo.fecha_inicio.strftime('%Y')
    # Crear la cadena de fecha_inicio del periodo en español
    inicio_mes_periodo = MONTHS_SPANISH.get(inicio_mes_periodo)
    periodo = f"{inicio_dia_periodo} de {inicio_mes_periodo} del {inicio_anio_periodo}"


    # Obtener la encuesta "REVIN025"
    encuesta = Encuesta.objects.get(nombre="REVIN025")

    # Obtener las secciones que pertenecen a esta encuesta
    secciones = Seccion.objects.filter(encuesta=encuesta).order_by('id')

    # Crear un archivo Excel en memoria
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fecha_actual = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = f'Reporte_REVIN026_{selprofe}_{fecha_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Escribir el DataFrame en un archivo Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Obtener la hoja de cálculo
        workbook = writer.book
        worksheet = workbook.create_sheet(title='REVIN026')
        # Insertar la imagen logo uts
        img_path = 'static/img/logo_uts.png'  # Reemplaza esto con la ruta a tu imagen
        img = Image(img_path)
        img.width = 180  # Ajusta el ancho de la imagen
        img.height = 119  # Ajusta la altura de la imagen
        worksheet.add_image(img, 'A1')  # A1 es la celda de inicio de la imagen

                # Función auxiliar para combinar celdas y ajustar texto
        def combinar_y_ajustar_texto(
                celdas, 
                texto, 
                font_size, 
                bold, 
                color='000000', 
                bg_color=None, 
                ajustar_ancho=False, 
                alineacion_horizontal='center', 
                alineacion_vertical='center', 
                wrap_text=False, 
                indent=0, 
                border=None,
            ):
            
            worksheet.merge_cells(celdas)
            cell = worksheet[celdas.split(':')[0]]
            cell.value = texto
            cell.alignment = Alignment(horizontal=alineacion_horizontal, vertical=alineacion_vertical, wrap_text=wrap_text, indent=indent)
            cell.font = Font(name='Arial Narrow', size=font_size, bold=bold, color=color)

            # Ajustar el ancho de la columna si se especifica
            if ajustar_ancho:
                # Obtener la letra de la columna
                col_letter = cell.column_letter
                
                # Calcular el ancho basado en la longitud del texto
                max_length = len(texto)
                adjusted_width = (max_length + 2) * (font_size / 10)  # Ajuste simple para el ancho basado en el tamaño de la fuente
                
                # Aplicar el ancho calculado a la columna
                worksheet.column_dimensions[col_letter].width = adjusted_width

            # Aplicar color de fondo si se especifica
            if bg_color:
                fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                for row_cells in worksheet[celdas]:
                    for cell in row_cells:
                        cell.fill = fill

            # Aplicar borde si se especifica
            if border:
                for row_cells in worksheet[celdas]:
                    for cell in row_cells:
                        cell.border = border

        thin_border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
                        
        # Ajustar los textos en las celdas
        combinar_y_ajustar_texto('C1:I2', "Universidad Tecnológica de Salamanca", 20, True, color='002078')
        combinar_y_ajustar_texto('C3:I3', "Organismo Público Descentralizado del Gobierno de Guanajuato", 14, False, color='002078')
        combinar_y_ajustar_texto('I6', "REVIN026-A", 12, True, ajustar_ancho=True)
        combinar_y_ajustar_texto('A7:I7', "RESULTADOS DE LA EVALUACIÓN DEL SERVICIO", 14, True)
        combinar_y_ajustar_texto('A8', "Nombre del Curso:", 12, True, ajustar_ancho=True, alineacion_horizontal='left')
        combinar_y_ajustar_texto('B8:E8', selcurso.nombre, 12, False, alineacion_horizontal='left')
        combinar_y_ajustar_texto('A9', "No. de Horas:", 12, True, ajustar_ancho=True, alineacion_horizontal='left')
        combinar_y_ajustar_texto('B9:E9', selcurso.duracion, 12, False, alineacion_horizontal='left')
        combinar_y_ajustar_texto('A10', "Instructor:", 12, True, ajustar_ancho=True, alineacion_horizontal='left')
        combinar_y_ajustar_texto('B10:E10', selprofe.nombre_completo(), 12, False, alineacion_horizontal='left')
        combinar_y_ajustar_texto('A11', "Fecha de Inicio:", 12, True, ajustar_ancho=True, alineacion_horizontal='left')
        combinar_y_ajustar_texto('B11:E11', periodo, 12, False, alineacion_horizontal='left')

        # Escribir secciones y preguntas
        row = 14
        texts = ["Excelente", "Muy bueno", "Bueno", "Regular", "Malo", "Muy malo", "TOTAL", "PROMEDIO"]
        promedios_generales = []

        for seccion in secciones:
            seccion_nombre = f"{seccion.id}. {seccion.nombre}"
            combinar_y_ajustar_texto(
                f'A{row}:A{row}', 
                seccion_nombre, 
                12, 
                True, 
                color='000000', 
                bg_color='c0c0c0', 
                ajustar_ancho=False, 
                alineacion_horizontal='left'
            )
            # Escribir los textos en las columnas B en adelante usando combinar_y_ajustar_texto
            for col, text in enumerate(texts, start=2):  # start=2 para comenzar desde la columna B
                col_letter = chr(64 + col)  # Convertir número de columna a letra (B, C, D, etc.)
                combinar_y_ajustar_texto(
                    f'{col_letter}{row}:{col_letter}{row}', 
                    text, 
                    12, 
                    True, 
                    color='000000', 
                    bg_color='c0c0c0', 
                    ajustar_ancho=True, 
                    alineacion_horizontal='center'
                )
            row += 1

            suma_promedios = 0
            total_preguntas = 0

            preguntas = Pregunta.objects.filter(seccion=seccion).order_by('id')
            start_row = row - 1  # Guardar la fila inicial para la sección

            for pregunta in preguntas:
                # Verificar si la pregunta es "Quejas, sugerencias o felicitaciones:"
                if pregunta.texto_pregunta == "Quejas, sugerencias o felicitaciones:":
                    continue  # Saltar al siguiente ciclo del bucle si la pregunta es la misma

                if type == "AC":   
                    respuestas = EncuestaAlumno.objects.filter(curso_alumno__profesor=selprofe, curso_alumno__curso=selcurso, curso_alumno__periodo=selperiodo, pregunta=pregunta)
                elif type == "AE":
                    respuestas = EncuestaEstudiante.objects.filter(curso_alumno__instructor=selprofe, curso_alumno__curso=selcurso, curso_alumno__periodo=selperiodo, pregunta=pregunta)

                if not respuestas.exists():
                    messages.error(request, 'No existen registros que cumplan con los filtros seleccionados.')
                    if type == "AC":
                        return redirect('/reportes-cele')
                    elif type == "AE":
                        return redirect('/reportes-edcon')

                if pregunta.texto_pregunta == "Recomendaría al instructor":
                    excelente = respuestas.filter(texto_respuesta="Si").count()
                else:
                    excelente = respuestas.filter(texto_respuesta=int(10)).count()

                muy_bueno = respuestas.filter(texto_respuesta=int(9)).count()
                bueno = respuestas.filter(texto_respuesta=int(8)).count()
                regular = respuestas.filter(texto_respuesta=int(7)).count()
                malo = respuestas.filter(texto_respuesta=int(6)).count()

                if pregunta.texto_pregunta == "Recomendaría al instructor":
                    muy_malo = respuestas.filter(texto_respuesta="No").count()
                else:
                    muy_malo = respuestas.filter(texto_respuesta=int(5)).count()

                total = excelente + muy_bueno + bueno + regular + malo + muy_malo
                promedio = round((excelente * 10 + muy_bueno * 9 + bueno * 8 + regular * 7 + malo * 6 + muy_malo * 5) / total, 2) if total > 0 else 0

                # Sumar el promedio de esta pregunta al total de promedios
                suma_promedios += promedio
                total_preguntas += 1

                combinar_y_ajustar_texto(f'B{row}', excelente, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'C{row}', muy_bueno, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'D{row}', bueno, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'E{row}', regular, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'F{row}', malo, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'G{row}', muy_malo, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'H{row}', total, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'I{row}', promedio, 10, False, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
                combinar_y_ajustar_texto(f'A{row}', pregunta.texto_pregunta, 10, False, color='000000', ajustar_ancho=True, alineacion_horizontal='left')
                row += 1

            # Calcular el Promedio General de la sección
            promedio_general = round(suma_promedios / total_preguntas, 2) if total_preguntas > 0 else 0
            promedios_generales.append(promedio_general)

            # Agregar la línea "Promedio General" debajo de la última pregunta
            combinar_y_ajustar_texto(f'A{row}', "Promedio General", 12, True, color='000000', ajustar_ancho=True, alineacion_horizontal='left')
            combinar_y_ajustar_texto(f'I{row}', promedio_general, 12, True, color='000000', ajustar_ancho=False, alineacion_horizontal='center')
            row += 2  # Mover a la siguiente fila después de agregar el "Promedio General"

            # Aplicar el borde a toda la sección
            end_row = row - 2  # Fila final para la sección
            worksheet.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            for r in range(start_row, end_row + 1):
                for c in range(1, 10):  # Ajusta el rango de columnas si es necesario (1 a 9 para A a I)
                    cell = worksheet.cell(row=r, column=c)
                    cell.border = thin_border

        combinar_y_ajustar_texto(f'A{row}:A{row}', "Calificación General del Curso:", 12, True, color='000000', bg_color='c0c0c0', ajustar_ancho=True, alineacion_horizontal='left', border=thin_border)
        combinar_y_ajustar_texto(f'B{row}:H{row}', "", 12, True, color='000000', bg_color='c0c0c0', border=thin_border)
        # Aplicar el borde a todas las celdas combinadas
        for col in range(ord('A'), ord('H') + 1):  # De la columna A a la H
            cell = worksheet[f'{chr(col)}{row}']
            cell.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
        calificacion_general = round(sum(promedios_generales) / len(secciones), 2) if secciones else 0
        combinar_y_ajustar_texto(f'I{row}:I{row}', calificacion_general, 12, True, color='000000', bg_color='c0c0c0', ajustar_ancho=False, alineacion_horizontal='center', border=thin_border)
        # Mover dos filas hacia abajo
        row += 2

        # Agregar contenido de "Quejas, sugerencias o felicitaciones:"
        for pregunta in preguntas:
            if pregunta.texto_pregunta == "Quejas, sugerencias o felicitaciones:":
                if type == "AC":   
                    respuestas = EncuestaAlumno.objects.filter(curso_alumno__profesor=selprofe, curso_alumno__curso=selcurso, curso_alumno__periodo=selperiodo, pregunta=pregunta)
                elif type == "AE":
                    respuestas = EncuestaEstudiante.objects.filter(curso_alumno__instructor=selprofe, curso_alumno__curso=selcurso, curso_alumno__periodo=selperiodo, pregunta=pregunta)

                quejas_sugerencias = " *".join([r.texto_respuesta for r in respuestas])

                # Agregar la pregunta en la columna A
                combinar_y_ajustar_texto(f'A{row}', pregunta.texto_pregunta, 12, True, color='000000', ajustar_ancho=True, alineacion_horizontal='left')

                # Mover a la siguiente fila para el contenido de las quejas
                row += 1

                # Agregar las quejas/sugerencias en la columna A (puedes ajustar la columna según tu diseño)
                combinar_y_ajustar_texto(
                    f'A{row}:I{row}', 
                    quejas_sugerencias, 
                    12, 
                    False, 
                    color='000000', 
                    bg_color='ffffff', 
                    ajustar_ancho=False, 
                    alineacion_horizontal='center', 
                    alineacion_vertical='center', 
                    wrap_text=True, 
                    indent=1, 
                    border=thin_border
                )


                # Mover a la siguiente fila después de agregar el contenido
                row += 1

                break  # Salir del bucle una vez encontrado el texto de la pregunta
        
        # Fila para "Firma del Instructor (Recibí):"
        row += 4

        # Combinar las celdas en el rango B{start_row}:E{end_row}
        worksheet.merge_cells(f'B{row}:E{row}')

        # Eliminar los bordes internos
        for row_num in range(row, row + 1):
            for col in ['B', 'C', 'D', 'E']:
                cell = worksheet[f'{col}{row_num}']
                cell.border = Border(
                    left=Side(border_style=None),
                    right=Side(border_style=None),
                    top=Side(border_style=None),
                    bottom=Side(border_style='medium', color='6D6E71')
                )
        combinar_y_ajustar_texto(f'A{row}', 'Firma del Instructor (Recibí):', 12, True, color='000000', ajustar_ancho=False, alineacion_horizontal='left')

        # Fila para "Fecha de Recibido:"
        row += 4

        # Combinar las celdas en el rango B{start_row}:E{end_row}
        worksheet.merge_cells(f'B{row}:E{row}')

        # Eliminar los bordes internos
        for row_num in range(row, row + 1):
            for col in ['B', 'C', 'D', 'E']:
                cell = worksheet[f'{col}{row_num}']
                cell.border = Border(
                    left=Side(border_style=None),
                    right=Side(border_style=None),
                    top=Side(border_style=None),
                    bottom=Side(border_style='medium', color='6D6E71')
                )
        combinar_y_ajustar_texto(f'A{row}', 'Fecha de Recibido:', 12, True, color='000000', ajustar_ancho=False, alineacion_horizontal='left')

    return response

class EncuestaAlumnoCeleListView(ListView):
    model = CursoAlumno
    context_object_name = 'encuestas'
    ordering = ['curso']
    paginate_by = 10

    def get_queryset(self):
        today = fecha_actual.today()
        queryset = CursoAlumno.objects.filter(alumno=self.request.user).order_by('-periodo__fecha_fin')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario_log = Alumno.objects.get(username=self.request.user.username)
        registros = RegistrosCELE.objects.all()
        
        context['status'] = 'alumno'
        context['usuario_log'] = usuario_log
        context['registros'] = registros
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Alumnos CELE":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(EncuestaAlumnoCeleListView, self).get(*args, **kwargs)

def is_valid_queryparam(param):
    return param != '' and param is not None
def invalid_queryparam(param):
    return param == '' and param is None


class SearchEncuestaAlumnoCeleView(ListView):
    model = CursoAlumno
    context_object_name = 'encuestas'      # default >> erf24/post_list.html
    ordering = ['curso']
    paginate_by = 10

    def get_queryset(self): # new
        search = self.request.GET.get('q')
        today = fecha_actual.today()

        if is_valid_queryparam(search):
            obj = CursoAlumno.objects.filter(Q(curso__nombre__icontains=search, alumno=self.request.user)).distinct().order_by('-periodo__fecha_fin')

        if invalid_queryparam(search):
            obj = CursoAlumno.objects.annotate(numero_de_alumnos=Count('alumno'))
    
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario_log = Alumno.objects.get(username=self.request.user.username)
        registros = RegistrosCELE.objects.all()
        
        context['status'] = 'alumno'
        context['usuario_log'] = usuario_log
        context['registros'] = registros
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Alumnos CELE":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(SearchEncuestaAlumnoCeleView, self).get(*args, **kwargs)

class EncuestaEstudianteEdconListView(ListView):
    model = CursoEstudiante
    context_object_name = 'encuestas'
    ordering = ['curso']
    paginate_by = 10

    def get_queryset(self):
        today = fecha_actual.today()
        queryset = CursoEstudiante.objects.filter(estudiante=self.request.user).order_by('-periodo__fecha_fin')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario_log = Estudiante.objects.get(username=self.request.user.username)
        registros = RegistrosEDCON.objects.all()
        
        context['status'] = 'estudiante'
        context['usuario_log'] = usuario_log
        context['registros'] = registros
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Estudiantes EDCON":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(EncuestaEstudianteEdconListView, self).get(*args, **kwargs)

def is_valid_queryparam(param):
    return param != '' and param is not None
def invalid_queryparam(param):
    return param == '' and param is None


class SearchEncuestaEstudianteEdconView(ListView):
    model = CursoEstudiante
    context_object_name = 'encuestas'      # default >> erf24/post_list.html
    ordering = ['curso']
    paginate_by = 10

    def get_queryset(self): # new
        search = self.request.GET.get('q')
        today = fecha_actual.today()

        if is_valid_queryparam(search):
            obj = CursoEstudiante.objects.filter(Q(curso__nombre__icontains=search, estudiante=self.request.user)).distinct().order_by('-periodo__fecha_fin')

        if invalid_queryparam(search):
            obj = CursoEstudiante.objects.annotate(numero_de_alumnos=Count('estudiante'))
    
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario_log = Estudiante.objects.get(username=self.request.user.username)
        registros = RegistrosEDCON.objects.all()
        
        context['status'] = 'estudiante'
        context['usuario_log'] = usuario_log
        context['registros'] = registros
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Estudiantes EDCON":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(SearchEncuestaEstudianteEdconView, self).get(*args, **kwargs)

class RegistroCeleListView(ListView):
    model = RegistrosCELE
    context_object_name = 'registros'
    ordering = ['curso_alumno']
    paginate_by = 10

    def get_queryset(self):
        today = fecha_actual.today()
        queryset = RegistrosCELE.objects.all().order_by('-curso_alumno__periodo__fecha_fin')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupos = self.request.user.groups.all()
        usuario = self.request.user
        usuario_log = usuario

        cursos = CursoCELE.objects.all()
        profesores = Profesor.objects.all()
        periodos = PeriodoCELE.objects.filter(activo=True)
        
        if usuario.is_superuser == 1:
            status = 'admin'

        for grupo in grupos:
            if grupo.name == 'Administradores CELE':
                status = 'admin'
            else:
                status = 'null'
        
        context['status'] = status
        context['usuario_log'] = usuario_log
        context['cursos'] = cursos
        context['profesores'] = profesores
        context['periodos'] = periodos

        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Administradores CELE":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(RegistroCeleListView, self).get(*args, **kwargs)

def is_valid_queryparam(param):
    return param != '' and param is not None
def invalid_queryparam(param):
    return param == '' and param is None



class SearchRegistroCeleView(ListView):
    model = RegistrosCELE
    context_object_name = 'registros'      # default >> erf24/post_list.html
    ordering = ['curso_alumno']
    paginate_by = 10

    def get_queryset(self): # new
        search = self.request.GET.get('q')
        today = fecha_actual.today()

        if is_valid_queryparam(search):
            obj = RegistrosCELE.objects.annotate(nombres=Concat('usuario__nombre', V(' '),'usuario__apellido_paterno', V(' '), 'usuario__apellido_materno')).filter(Q(curso_alumno__curso__nombre__icontains=search) | Q(nombres__icontains=search)).distinct().order_by('-curso_alumno__periodo__fecha_fin')

        if invalid_queryparam(search):
            obj = RegistrosCELE.objects.annotate(numero_de_alumnos=Count('usuario'))
    
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupos = self.request.user.groups.all()
        usuario = self.request.user
        usuario_log = usuario

        cursos = CursoCELE.objects.all()
        profesores = Profesor.objects.all()
        periodos = PeriodoCELE.objects.filter(activo=True)

        if usuario.is_superuser == 1:
            status = 'admin'

        for grupo in grupos:
            if grupo.name == 'Administradores CELE':
                status = 'admin'
            else:
                status = 'null'
        
        context['status'] = status
        context['usuario_log'] = usuario_log
        context['cursos'] = cursos
        context['profesores'] = profesores
        context['periodos'] = periodos
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Administradores CELE":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(SearchRegistroCeleView, self).get(*args, **kwargs)



class RegistroEdconListView(ListView):
    model = RegistrosEDCON
    context_object_name = 'registros'
    ordering = ['curso_alumno']
    paginate_by = 10

    def get_queryset(self):
        today = fecha_actual.today()
        queryset = RegistrosEDCON.objects.all().order_by('-curso_alumno__periodo__fecha_fin')
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupos = self.request.user.groups.all()
        usuario = self.request.user
        usuario_log = usuario

        cursos = CursoEDCON.objects.all()
        profesores = Instructor.objects.all()
        periodos = PeriodoEDCON.objects.filter(activo=True)

        if usuario.is_superuser == 1:
            status = 'admin'

        for grupo in grupos:
            if grupo.name == 'Administradores EDCON':
                status = 'admin'
            else:
                status = 'null'
        
        context['status'] = status
        context['usuario_log'] = usuario_log
        context['cursos'] = cursos
        context['profesores'] = profesores
        context['periodos'] = periodos
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Administradores EDCON":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(RegistroEdconListView, self).get(*args, **kwargs)

def is_valid_queryparam(param):
    return param != '' and param is not None
def invalid_queryparam(param):
    return param == '' and param is None



class SearchRegistroEdconView(ListView):
    model = RegistrosEDCON
    context_object_name = 'registros'      # default >> erf24/post_list.html
    ordering = ['curso_alumno']
    paginate_by = 10

    def get_queryset(self): # new
        search = self.request.GET.get('q')
        today = fecha_actual.today()

        if is_valid_queryparam(search):
            obj = RegistrosEDCON.objects.annotate(nombres=Concat('usuario__nombre', V(' '),'usuario__apellido_paterno', V(' '), 'usuario__apellido_materno')).filter(Q(curso_alumno__curso__nombre__icontains=search) | Q(nombres__icontains=search)).distinct().order_by('-curso_alumno__periodo__fecha_fin')

        if invalid_queryparam(search):
            obj = RegistrosEDCON.objects.annotate(numero_de_alumnos=Count('usuario'))
    
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grupos = self.request.user.groups.all()
        usuario = self.request.user
        usuario_log = usuario

        cursos = CursoEDCON.objects.all()
        profesores = Instructor.objects.all()
        periodos = PeriodoEDCON.objects.filter(activo=True)

        if usuario.is_superuser == 1:
            status = 'admin'

        for grupo in grupos:
            if grupo.name == 'Administradores EDCON':
                status = 'admin'
            else:
                status = 'null'
        
        context['status'] = status
        context['usuario_log'] = usuario_log
        context['cursos'] = cursos
        context['profesores'] = profesores
        context['periodos'] = periodos
        return context
    
    def get(self, *args, **kwargs):
        grupos = self.request.user.groups.all()
        for grupo in grupos:
            if grupo.name != "Administradores EDCON":
                return redirect('/no_autorizado')
        if not self.request.user.is_authenticated:
            return redirect('/login')
        return super(SearchRegistroEdconView, self).get(*args, **kwargs)